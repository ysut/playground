import openpyxl
import tqdm

def combine(excel_files: list, sheet_names: list):
    master_wb = openpyxl.Workbook()
    dummy_ws = master_wb.active
    master_wb.remove(dummy_ws)
    for i, (excel_file, sheet_name) in enumerate(zip(excel_files, sheet_names)):
        tmp_wb = openpyxl.load_workbook(filename=excel_file)
        tmp_ws = tmp_wb.worksheets[0]
        master_wb.create_sheet(index=i, title=sheet_name)
        master_ws = master_wb.worksheets[i]
        for row in tmp_ws:
            for cell in row:
                master_ws[cell.coordinate].value = cell.value
    return master_wb



