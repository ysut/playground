import openpyxl
from openpyxl.styles import Font
import pathlib
from liftover import get_lifter

# Choose one page from 'assessment-tools', 'acmg-classification', 'publications', and 'gene-assessment'
franklin_page = 'assessment-tools'  
ucsc_width: int = 30 # Display the range specified plus or minus bp


class HyperLinks:
    def __init__(self, input_excel_file, hgmd_link, ucsc_link, franklin_link, decipher_link, symbol_col):
        print('Inserting hyperlinks ......')
        self.input_file = input_excel_file
        self.workbook = openpyxl.load_workbook(self.input_file)
        self.sheet_names = self.workbook.sheetnames
        self.sheets = self.workbook.worksheets
        self.symbol_col = symbol_col
        self.insert_col_hgmd = openpyxl.utils.cell.get_column_letter(int(hgmd_link))
        self.insert_col_ucsc = openpyxl.utils.cell.get_column_letter(int(ucsc_link))
        self.insert_col_frnk = openpyxl.utils.cell.get_column_letter(int(franklin_link)) 
        self.insert_col_dcpr = openpyxl.utils.cell.get_column_letter(int(decipher_link))
    

    def __del__(self):
        self.workbook.close()


    def _search_col_letter(self, keywords):
        result = {}
        for i, sheet in enumerate(self.sheets):
            key_letter_dict = {}
            for cell in sheet['1']:
                try:
                    value = str(cell.value)
                except: 
                    continue
                for keyword in keywords:
                    if value == keyword:
                        letter = cell.column_letter
                        key_letter_dict[keyword] = letter
                        break
                    else:
                        pass
            result[self.sheet_names[i]] = key_letter_dict
        return result 
    

    def format_hyperlink():
        pass


    def insert_hyperlink_cols(self):
        for i, sheet in enumerate(self.sheets):
            if self.sheet_names[i] != 'Count' and self.sheet_names[i] != 'XHMM':
                if sheet.max_row > 1:
                    sheet.insert_cols(1)
                    sheet.insert_cols(2)
                    sheet.insert_cols(3)
                    sheet.insert_cols(4)
                    for i in range(sheet.max_row):
                        sheet[f'A{i+1}'] = 'HGMD'
                        sheet[f'B{i+1}'] = 'UCSC'
                        sheet[f'C{i+1}'] = 'Franklin'
                        sheet[f'D{i+1}'] = 'DECIPHER'
                else:
                    pass
            else:
                pass
    

    def hyperlink_hgmd(self):
        col_letters = self._search_col_letter([self.symbol_col])
        for sheet_num, sheet in enumerate(self.sheets):
            try:
                symbol_col_letter = col_letters[self.sheet_names[sheet_num]][self.symbol_col]
            except KeyError:
                continue
            symbol_col_letter = col_letters[self.sheet_names[sheet_num]][self.symbol_col]
            max_row = sheet.max_row
            for index in range(max_row):
                if index != 0:
                    gene_symbol = sheet[f'{symbol_col_letter}{index + 1}'].value
                    hgmd_url = f'https://my.qiagendigitalinsights.com/bbp/view/hgmd/pro/gene.php?gene={gene_symbol}'
                    sheet[f'{self.insert_col_hgmd}{index + 1}'].hyperlink = hgmd_url
                else:
                    pass

        
    def hyperlink_ucsc(self):
        col_letters = self._search_col_letter([self.symbol_col, 'CHROM', 'Start', 'End'])
        for sheet_num, sheet in enumerate(self.sheets):
            try:
                symbol_col_letter = col_letters[self.sheet_names[sheet_num]][self.symbol_col]
            except KeyError:
                continue
            # symbol_col_letter = col_letters[self.sheet_names[sheet_num]][self.symbol_col]
            chrom_col_letter = col_letters[self.sheet_names[sheet_num]]['CHROM']
            start_col_letter = col_letters[self.sheet_names[sheet_num]]['Start']
            end_col_letter = col_letters[self.sheet_names[sheet_num]]['End']
            max_row = sheet.max_row
            for index in range(max_row):
                if index != 0:
                    chrom = sheet[f'{chrom_col_letter}{index + 1}'].value
                    start = int(sheet[f'{start_col_letter}{index + 1}'].value)
                    end = int(sheet[f'{end_col_letter}{index + 1}'].value)
                    ucsc_url = f'https://genome-asia.ucsc.edu/cgi-bin/hgTracks?db=hg19&highlight=hg19.chr{chrom}%3A{start}-{end}&position=chr{chrom}%3A{start - ucsc_width}-{end + ucsc_width}'
                    sheet[f'{self.insert_col_ucsc}{index + 1}'].hyperlink = ucsc_url
                else:
                    pass

    def hyperlink_franklin(self):
        col_letters = self._search_col_letter([self.symbol_col, 'CHROM', 'POS', 'REF', 'ALT'])
        for sheet_num, sheet in enumerate(self.sheets):
            try:
                symbol_col_letter = col_letters[self.sheet_names[sheet_num]][self.symbol_col]
            except KeyError:
                continue
            # symbol_col_letter = col_letters[self.sheet_names[sheet_num]][self.symbol_col]
            chrom_col_letter = col_letters[self.sheet_names[sheet_num]]['CHROM']
            pos_col_letter = col_letters[self.sheet_names[sheet_num]]['POS']
            ref_col_letter = col_letters[self.sheet_names[sheet_num]]['REF']
            alt_col_letter = col_letters[self.sheet_names[sheet_num]]['ALT']
            max_row = sheet.max_row
            for index in range(max_row):
                if index != 0:
                    gene_symbol = sheet[f'{symbol_col_letter}{index + 1}'].value
                    chrom = sheet[f'{chrom_col_letter}{index + 1}'].value
                    pos = int(sheet[f'{pos_col_letter}{index + 1}'].value)
                    ref = sheet[f'{ref_col_letter}{index + 1}'].value
                    alt = sheet[f'{alt_col_letter}{index + 1}'].value
                    frnk_url = f'https://franklin.genoox.com/clinical-db/variant/snp/chr{chrom}-{pos}-{ref}-{alt}?app={franklin_page}'
                    sheet[f'{self.insert_col_frnk}{index + 1}'].hyperlink = frnk_url
                else:
                    pass

    def hyperlink_decipher(self):
        col_letters = self._search_col_letter([self.symbol_col, 'CHROM', 'POS', 'REF', 'ALT'])
        for sheet_num, sheet in enumerate(self.sheets):
            try:
                symbol_col_letter = col_letters[self.sheet_names[sheet_num]][self.symbol_col]
            except KeyError:
                continue
            chrom_col_letter = col_letters[self.sheet_names[sheet_num]]['CHROM']
            pos_col_letter = col_letters[self.sheet_names[sheet_num]]['POS']
            ref_col_letter = col_letters[self.sheet_names[sheet_num]]['REF']
            alt_col_letter = col_letters[self.sheet_names[sheet_num]]['ALT']
            max_row = sheet.max_row
            for index in range(max_row):
                if index != 0:
                    gene_symbol = sheet[f'{symbol_col_letter}{index + 1}'].value
                    chrom = sheet[f'{chrom_col_letter}{index + 1}'].value
                    pos = int(sheet[f'{pos_col_letter}{index + 1}'].value)
                    ref = sheet[f'{ref_col_letter}{index + 1}'].value
                    alt = sheet[f'{alt_col_letter}{index + 1}'].value
                    converter = get_lifter('hg19', 'hg38')
                    try: 
                        pos_hg38 = converter[chrom][pos][0][1]
                    except IndexError:
                        dcpr_url_gene = f'https://www.deciphergenomics.org/gene/{gene_symbol}/overview/protein-genomic-info'
                        sheet[f'{self.insert_col_dcpr}{index + 1}'].hyperlink = dcpr_url_gene
                        continue
                    dcpr_url = f'https://www.deciphergenomics.org/sequence-variant/{chrom}-{pos_hg38}-{ref}-{alt}/genes/{gene_symbol}/protein-genomic-info'
                    sheet[f'{self.insert_col_dcpr}{index + 1}'].hyperlink = dcpr_url
                else:
                    pass 
    
    # def hyperlink_decipher_cnv(self):
    #     col_letters = self._search_col_letter([self.symbol_col, 'CHROM', 'POS', 'REF', 'ALT'])
    #     for sheet_num, sheet in enumerate(self.sheets):
    #         try:
    #             symbol_col_letter = col_letters[self.sheet_names[sheet_num]][self.symbol_col]
    #         except KeyError:
    #             continue
    #         chrom_col_letter = col_letters[self.sheet_names[sheet_num]]['CHROM']
    #         pos_col_letter = col_letters[self.sheet_names[sheet_num]]['POS']
    #         ref_col_letter = col_letters[self.sheet_names[sheet_num]]['REF']
    #         alt_col_letter = col_letters[self.sheet_names[sheet_num]]['ALT']
    #         max_row = sheet.max_row

    #     pass


    def _save(self):
        input_path = pathlib.Path(self.input_file)
        input_file_name = input_path.stem
        input_file_dir = input_path.parent
        output = f'{input_file_dir}/{input_file_name}.weslinked.xlsx'
        print(f'Creating {output} ......')
        self.workbook.save(output)

    def all_links(self):
        self.insert_hyperlink_cols()
        print('HGMD links ......')
        self.hyperlink_hgmd()
        print('UCSC links ......')
        self.hyperlink_ucsc()
        print('Franklin links ......')
        self.hyperlink_franklin()
        print('DECIPHER links ......')
        self.hyperlink_decipher()

        self._save()
        self.__del__()

# class HyperLinks(ExcelEdit):
#     def __init__(self, dataframe, output_dir, gene_symbol_col):
#         super().__init__(dataframe, output_dir)
#         self.gene_symobl_col = gene_symbol_col

#     def hyperlink_hgmd(self, worksheet, link_col):
#         col_letters_dict = super().search_col_letter('1', ['Gene.refGene', link_col])
# #       super().search_col_region()
#         gene_symbol_col_letter = col_letters_dict['Gene.refGene']
#         insert_link_col_letter = col_letters_dict[link_col]
#         for index in range(self.max_row):
#             gene_symbol = worksheet[f'{gene_symbol_col_letter}{index + 1}'].value
#             hgmd_url = f'https://my.qiagendigitalinsights.com/bbp/view/hgmd/pro/gene.php?gene={gene_symbol}'
#             worksheet[f'{insert_link_col_letter}{index + 1}'].hyperlink = hgmd_url

#     # def hyperlink_decipher(self, worksheet, link_col):
#     #     gene_symbol_col_letter = col_letters_dict['Gene.refGene']
#     #     chrom_col_letter = col_letters_dict['CHROM']
#     #     pos_col_letter = col_letters_dict['POS']
#     #     ref_col_letter = col_letters_dict['REF']
#     #     alt_col_letter = col_letters_dict['ALT']
#     #     insert_link_col_letter = col_letters_dict[link_col]

#     #     for index in range(len(df)):
#     #         if index != 0:
#     #             chrom = worksheet[f'{chrom_col_letter}{index + 1}'].value
#     #             pos = int(worksheet[f'{pos_col_letter}{index + 1}'].value)
#     #             ref = worksheet[f'{ref_col_letter}{index + 1}'].value
#     #             alt = worksheet[f'{alt_col_letter}{index + 1}'].value
#     #             converter = get_lifter('hg19', 'hg38')
#     #             pos_hg38 = converter[chrom][pos][0][1]
#     #             insert_link_col_letter = col_letters_dict[link_col]
#     #             gene_symbol = ws[f'{gene_symbol_col_letter}{index + 1}'].value
#     #             decipher_url = f'https://www.deciphergenomics.org/sequence-variant/{chrom}-{pos_hg38}-{ref}-{alt}/genes/{gene_symbol}/protein-genomic-info'
#     #             worksheet[f'{insert_link_col_letter }{index + 1}'].hyperlink = decipher_url
#     #         else:
#     #             pass