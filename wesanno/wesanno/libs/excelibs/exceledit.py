import datetime
import pandas as pd
import openpyxl as px
import gzip
import re

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScale, IconSet
from openpyxl.formatting.rule import FormatObject, Rule
from openpyxl.styles import Font
from openpyxl.styles.differential import DifferentialStyle

from liftover import get_lifter


class ExcelEdit:
    def __init__(self, dataframe, output_dir):
        self.dataframe = dataframe
        self.max_row = len(dataframe)
        self.output_dir = output_dir
        self.dt_now = datetime.datetime.now()
        timestamp = self.dt_now.strftime('%Y%m%d-%H%M%S')
        dataframe.to_excel(f'{output_dir}/interval_{timestamp}.xlsx', index=False)
        self.workbook = px.load_workbook(f'{output_dir}/interval_{timestamp}.xlsx')
        self.worksheet = self.workbook.active

    def search_column_letter(self, row_num, keywords):
        result = {}
        for cell in self.worksheet[row_num]:
            try:
                value = str(cell.value)
            except: 
                continue
            for keyword in keywords:
                if value == keyword:
                    column_letter = cell.column_letter
                    result[keyword] = column_letter
                    break
                else:
                    pass
        return result

    def search_column_region(self):
        column_letters_dictionary = self.search_column_letter('1', list(self.dataframe.columns))
        result = {}
        for k, v in column_letters_dictionary.items():
            result[k] = f'{v}1:{v}{self.max_row}'   
        return result
    
    def overall_font_setting(self, font, size):
        overall_font = Font(name=font, size=size)
        for row in self.worksheet:
            for cell in row:
                self.worksheet[cell.coordinate].font = overall_font
    
    def final_output(self):
        self.overall_font_setting('Arial', 10)
        out_timestamp = self.dt_now.strftime('%Y%m%d-%H%M%S')
        self.workbook.save(f'{self.output_dir}/results_{out_timestamp}.xlsx')  

    
        
        