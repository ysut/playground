import os
import re
import urllib.parse
from dataclasses import dataclass
from logging import getLogger

import numpy as np
import openpyxl as xl
from openpyxl.styles import Font, Color
from tqdm import tqdm
import pandas as pd
from pandarallel import pandarallel

pandarallel.initialize(
    progress_bar=False, verbose=1, use_memory_fs=False,
    nb_workers=int(os.cpu_count() -1))

from logging import getLogger
logger = getLogger(__name__)

@dataclass
class ExcelSheetsDF:
    sheets: dict


class Hyperlink:
    def __init__(
            self, 
            gene_symbol_col: str,
            alt_col: str,
            assembly: str,
            pos19: str,
            pos38: str, 
            franklin_page: str, 
            ucsc_width: int,
            skip_sites: list,
            spliceai_maskFlag: bool,
            spliceai_dist: int,
            windowsFlag: bool
            ):
        # Set logger    
        self.logger = getLogger(__name__)

        # Configure arguments for hyperlink insertion
        self.gene_symbol_col: str = gene_symbol_col
        self.alt_col: str = alt_col
        self.assembly: str = assembly
        self.pos38: str = pos38
        self.franklin_page: str = franklin_page
        self.ucsc_width: int = ucsc_width
        self.skip_sites: list = skip_sites
        self.is_splai_masked: bool = spliceai_maskFlag
        self.splai_dist: int = spliceai_dist
        self.is_windows: bool = windowsFlag

        # Set query position
        if ((self.assembly == 'hg19') | (self.assembly == 'GRCh37')):
            self.query_pos_col: str = pos19
        elif ((self.assembly == 'hg38') | (self.assembly == 'GRCh38')):
            self.query_pos_col: str = pos38
        else:
            raise ValueError(f"Invalid assembly: {self.assembly}")
        self.logger.debug(f"Query position column: {self.query_pos_col}")


    ## Check gene symbol
    def __get_query_gene(self, gene_info) -> str:
        """ 
        When multiple genes are listed (e.g. overlaping regions), 
        use a first gene symbol.
        * Gene symbols should be separated by comma, semicolon, 
          colon, pipe, or slash.
        """
        try:
            candidate_genes = re.split('[,;:|/]', gene_info)
        except AttributeError:
            return gene_info
        else:
            return candidate_genes[0]


    ## Covert chromosome format (e.g. chr1 -> 1)
    def __convert_chrom_format(self, chrom) -> str:
        if str(chrom).startswith('chr'):
            return str(chrom).replace('chr', '')
        else:
            return str(chrom)


    ## Reorder columns
    def __rearrange_cols(self, df: pd.DataFrame) -> pd.DataFrame:
        lst: list = df.columns.tolist()
        specified_cols = [
            'HGMD', 'DECIPHER', 'SpliceAI_Lookup', 'UCSC', 'Franklin'
            ]
        
        # Count the number of inserted columns
        inserted_col: int = 0
        for specified_col in specified_cols:
            if specified_col in lst:
                inserted_col += 1

        # Get hyperlink columns
        hyperlink_cols: list = lst[-inserted_col:] 
        logger.info(f"Hyperlink columns: {hyperlink_cols}")
        # Get the rest of columns
        non_hyperlink_cols: list = lst[:-inserted_col]
        # Reorder columns
        rearranged: list = hyperlink_cols + non_hyperlink_cols
        rearranged_df = df[rearranged]
    
        return rearranged_df
    

    #1 HGMD
    def __generate_hgmd_url(self, row):        
        base_url = f"https://my.qiagendigitalinsights.com/bbp/view/hgmd/pro/gene.php"
        url: str = (
            f"{base_url}"
            f"?gene={self.__get_query_gene(row[self.gene_symbol_col])}"
            )

        return urllib.parse.quote(url)
    
    #2 UCSC
    def __generate_ucsc_url(self, row):
        base_url = "https://genome-asia.ucsc.edu/cgi-bin/hgTracks"
        start: int = int(row[self.query_pos_col])
        end: int = start + int(np.abs(len(row['REF']) - len(row[self.alt_col])))
        
        # Convert chromosome name (e.g. chr1 -> 1)
        chrom: str = self.__convert_chrom_format(row['CHROM'])

        url: str = (
            f"{base_url}"
            f"?db={self.assembly}"
            f"&highlight={self.assembly}.chr{chrom}:{start}-{end}"
            f"&position=chr{chrom}:{start - self.ucsc_width}-{end + self.ucsc_width}"
            )

        return urllib.parse.quote(url) 
    
    #3 Franklin
    def __generate_franklin_url(self, row):
        base_url = "https://franklin.genoox.com/clinical-db/variant"
        common_url: str = (
            f"{base_url}/snp/"
            f"{row['CHROM']}-{row[self.query_pos_col]}-{row['REF']}-{row[self.alt_col]}"
            )
        
        if self.assembly == 'hg19':
            url: str = f"{common_url}?app={self.franklin_page}"
        elif self.assembly == 'hg38':
            url: str = f"{common_url}-hg38?app={self.franklin_page}"

        return urllib.parse.quote(url)
    
    #4 DECIPHER
    def __generate_decipher_url(self, row):
        # Check gene symbol
        base_url = "https://www.deciphergenomics.org"
        query_gene: str = self.__get_query_gene(row[self.gene_symbol_col])
        
        # Convert chromosome name (e.g. chr1 -> 1)
        chrom: str = self.__convert_chrom_format(row['CHROM'])

        try: 
            query_pos38_col = int(row[self.pos38])
        except ValueError:
            url: str = (
                f"{base_url}/gene/"
                f"{query_gene}/overview/protein-genomic-info"
                )
        else:
            url: str = (
                f"{base_url}/sequence-variant/"
                f"{chrom}-{query_pos38_col}-{row['REF']}-{row[self.alt_col]}/"
                f"genes/{query_gene}/protein-genomic-info"
                )

        return urllib.parse.quote(url)

    #5 SpliceAI lookup
    def __generate_spliceailookup_url(self, row) -> urllib.parse.quote:
        if ((self.assembly == 'hg19') | (self.assembly == 'GRCh37')):
            query_genome_version: str = '37'
        elif ((self.assembly == 'hg38') | (self.assembly == 'GRCh38')):
            query_genome_version: str = '38'
        else:
            raise ValueError(f"Invalid assembly: {self.assembly}")
        
        base_url = "https://spliceailookup.broadinstitute.org"
        query_variant = (f"{row['CHROM']}-{row[self.query_pos_col]}"
                         f"-{row['REF']}-{row[self.alt_col]}")
        
        if self.is_splai_masked:
            mask = 1
        else:
            mask = 0
        
        url: str = (
            f"{base_url}/"
            f"#variant={query_variant}"
            f"&hg={query_genome_version}"
            f"&distance={self.splai_dist}&mask={mask}&ra=0"
            )

        if self.is_windows:
            return urllib.parse.quote(url, safe='#=&:/')
        else:
            return urllib.parse.quote(url)
    
    def insert_urls(self, df: pd.DataFrame) -> pd.DataFrame:
        # Insert hyperlinks
        if 'HGMD' not in self.skip_sites:
            if self.gene_symbol_col in df.columns:
                df['HGMD'] = df.parallel_apply(
                    self.__generate_hgmd_url, axis=1)
            else:
                logger.warning("Skip HGMD URL insertion.")
        else:
            self.logger.info("Skip HGMD URL insertion")
        
        if 'Franklin' not in self.skip_sites:
            df['Franklin'] = df.parallel_apply(
                self.__generate_franklin_url, axis=1)
        else:
            self.logger.info("Skip Franklin URL insertion")
        
        if 'DECIPHER' not in self.skip_sites:
            if self.gene_symbol_col in df.columns:
                df['DECIPHER'] = df.parallel_apply(
                    self.__generate_decipher_url, axis=1)
            else:
                logger.warning("Skip DECIPHER URL insertion.")
        else:
            self.logger.info("Skip DECIPHER URL insertion")

        if 'SpliceAI' not in self.skip_sites:
            df['SpliceAI_Lookup'] = df.parallel_apply(
                self.__generate_spliceailookup_url, axis=1)
        else:
            self.logger.info("Skip SpliceAI lookup URL insertion")

        if 'UCSC' not in self.skip_sites:
            df['UCSC'] = df.parallel_apply(
                self.__generate_ucsc_url, axis=1)
        else:
            self.logger.info("Skip UCSC URL insertion")

        # Reorder columns
        df = self.__rearrange_cols(df)

        return df

def load_excel_as_dataclass(input_execl) -> ExcelSheetsDF:
    exl = pd.ExcelFile(input_execl)       # Load input excel file
    sheet_names: list = exl.sheet_names   # Get all sheet names
    sheets: dict = {
        sheet_name: exl.parse(sheet_name) for sheet_name in sheet_names
        }

    return ExcelSheetsDF(sheets=sheets)


def generate_anno_sheets_list(input_sheets: list, skip_sheets: list) -> list:    
    anno_sheets: list = [
        sheet_name for sheet_name in input_sheets 
        if sheet_name not in set(skip_sheets)]

    return anno_sheets


def df_to_excel(dfs: ExcelSheetsDF, output_xlsx) -> None:
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        for sheet in dfs.sheets:
            dfs.sheets[sheet].to_excel(writer, sheet_name=sheet, index=False)


def __count_inserted_cols(sheet: xl.worksheet.worksheet.Worksheet) -> int:
    sites = ['HGMD', 'Franklin', 'DECIPHER', 'SpliceAI_Lookup', 'UCSC']
    inseted_col: int = 0
    for i in range(1,6):
        if sheet.cell(1, i).value in sites:
            inseted_col += 1
        else:
            pass
    
    return inseted_col


def convert_url_to_hyperlink(
        input_excel: str, anno_sheets: list, skip_sites: list
        ) -> xl.workbook.workbook.Workbook:
    
    sites = ['HGMD', 'Franklin', 'DECIPHER', 'SpliceAI_Lookup', 'UCSC']
    # max_col: int = 5 - len(skip_sites)
    wb = xl.load_workbook(input_excel)
    sheets = wb.worksheets

    for sheet in tqdm(sheets):
        if sheet.title in anno_sheets:
            inseted_col: int = __count_inserted_cols(sheet)            
            for col in sheet.iter_cols(min_row=2, max_row=sheet.max_row, 
                                    min_col=1, max_col=inseted_col):
                for cell in col:
                    col_letter = xl.utils.get_column_letter(cell.column)
                    cell.hyperlink = cell.value
                    cell.value = sheet[f"{col_letter}1"].value
                    cell.font = Font(
                        color=Color(rgb='FF0000FF'), underline='single'
                        )
        else:
            pass
    
    return wb


