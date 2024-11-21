import datetime
import pandas as pd
import openpyxl
import numpy as np
import pathlib

from openpyxl.formatting.rule import DataBar, ColorScale, IconSet, FormatObject
from openpyxl.formatting.rule import Rule
from openpyxl.formatting.rule import ColorScaleRule, IconSetRule
from openpyxl.styles import Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation

import toml
from liftover import get_lifter

class ExcelFormat():
    def __init__(self, input_excel_file):
        print('Loading a Excel file......')
        self.input_file = input_excel_file
        self.workbook = openpyxl.load_workbook(self.input_file)
        self.sheet_names = self.workbook.sheetnames
        self.sheets = self.workbook.worksheets
        pass


    def _search_col_letter(self, *args):
        self.col_region_dict = {}
        for i, sheet in enumerate(self.sheets):
            key_letter_dict = {}
            for cell in sheet['1']:
                try:
                    value = str(cell.value)
                except: 
                    continue
                for keyword in args:
                    if value == keyword:
                        letter = cell.column_letter
                        key_letter_dict[keyword] = letter
                        break
                    else:
                        pass
            self.col_region_dict[self.sheet_names[i]] = key_letter_dict
        return self.col_region_dict
    

    def insert_comment_cols(self):
        font_b = Font(bold=True, color='009688')
        font_lb = Font(bold=True, color='80CBC4')
        font_vus = Font(bold=True, color='607D8B')
        font_lp = Font(bold=True, color='FFA000')
        font_p = Font(bold=True, color='E85A70')
        dxf_b = DifferentialStyle(font=font_b)
        dxf_lb = DifferentialStyle(font=font_lb)
        dxf_vus = DifferentialStyle(font=font_vus)
        dxf_lp = DifferentialStyle(font=font_lp)
        dxf_p = DifferentialStyle(font=font_p)
        rule_b = Rule(type='cellIs', operator='equal',formula=['"Benign"'], dxf=dxf_b)
        rule_lb = Rule(type='cellIs', operator='equal',formula=['"Likely Benign"'], dxf=dxf_lb)
        rule_vus = Rule(type='cellIs', operator='equal',formula=['"VUS"'], dxf=dxf_vus)
        rule_lp = Rule(type='cellIs', operator='equal',formula=['"Likely Pathogenic"'], dxf=dxf_lp)
        rule_p = Rule(type='cellIs', operator='equal',formula=['"Pathogenic"'], dxf=dxf_p)

        for i, sheet in enumerate(self.sheets):
            if self.sheet_names[i] != 'Count':
                if sheet.max_row > 1:
                    dv = DataValidation(
                            type="list", 
                            formula1='"Benign,Likely Benign,VUS,Likely Pathogenic,Pathogenic"',
                            allow_blank=True,
                            showErrorMessage=True,
                            errorStyle="warning",
                            errorTitle="Are there any fields you want to enter?",
                            error="Would you like to enter any value?"
                            )
                    sheet.insert_cols(1)
                    sheet.insert_cols(2)
                    sheet['A1'] = 'Assessment'
                    sheet['B1'] = 'Comment'
                    dv.add(f'A2:A{sheet.max_row}')
                    sheet.add_data_validation(dv)
                    sheet.conditional_formatting.add(f'A2:A{sheet.max_row}', rule_b)
                    sheet.conditional_formatting.add(f'A2:A{sheet.max_row}', rule_lb)
                    sheet.conditional_formatting.add(f'A2:A{sheet.max_row}', rule_vus)
                    sheet.conditional_formatting.add(f'A2:A{sheet.max_row}', rule_lp)
                    sheet.conditional_formatting.add(f'A2:A{sheet.max_row}', rule_p)
                else:
                    pass
            else:
                pass

    def insert_hyperlink_cols(self):
        for i, sheet in enumerate(self.sheets):
            if sheet.max_row > 1:
                
                sheet.insert_cols(3) # HGMD
                sheet.insert_cols(4) # DECIPHER
                sheet.insert_cols(5) # Franklin 
                sheet.insert_cols(6) # SpliceAI lookup
                sheet.insert_cols(7) # UCSC
                sheet['C1'] = 'HGMD'
                sheet['D1'] = 'DECIPHER'
                sheet['E1'] = 'Franklin'
                sheet['F1'] = 'SpliceAI_lookup'
                sheet['G1'] = 'UCSC'

                for row in sheet.iter_rows(
                    min_row=2, max_row=sheet.max_row, min_col=3, max_col=7):
                    
                    for cell in row:
                        cell.hyperlink = f'{self.sheet_names[i]}!A{cell.row}'
                        cell.value = 'Link'

            else: # If brank sheet, do not insert hyperlink cols.
                pass


    # def coloring_hyperlink(self):
    #     for i, sheet in enumerate(self.sheets):
    #         if self.sheet_names[i] != 'Count':
    #             for col in sheet.
    #             if sheet.max_row > 1:


    #                 dv = DataValidation(
    #                         type="list", 
    #                         formula1='"Benign,Likely Benign,VUS,Likely Pathogenic,Pathogenic"',
    #                         allow_blank=True,
    #                         showErrorMessage=True,
    #                         errorStyle="warning",
    #                         errorTitle="Are there any fields you want to enter?",
    #                         error="Would you like to enter any value?"
    #                         )
    #                 sheet.insert_cols(1)
    #                 sheet.insert_cols(2)
    #                 sheet['A1'] = 'Assessment'
    #                 sheet['B1'] = 'Comment'
    #                 dv.add(f'A2:A{sheet.max_row}')
    #                 sheet.add_data_validation(dv)
    #                 sheet.conditional_formatting.add(f'A2:A{sheet.max_row}', rule_b)
    #                 sheet.conditional_formatting.add(f'A2:A{sheet.max_row}', rule_lb)
    #                 sheet.conditional_formatting.add(f'A2:A{sheet.max_row}', rule_vus)
    #                 sheet.conditional_formatting.add(f'A2:A{sheet.max_row}', rule_lp)
    #                 sheet.conditional_formatting.add(f'A2:A{sheet.max_row}', rule_p)
    #             else:
    #                 pass
    #         else:
    #             pass

    #                         font = Font(color='0000ee', underline='single')

    #                         #     sheet['A3'].font = font
    #                 # # sheet[f'A2:A{sheet.max_row}'].font = font
    #                 # # sheet[f'B3'].font = font
    #                 # # sheet[f'C3'].font = font
    #                 # # sheet[f'D5'].font = font




    def clinsig_color():
        pass
    def cadd_color():
        pass
    def sift_color():
        pass
    def pp2_color():
        pass
    def vest3_color():
        pass
    def dann_color():
        pass

    def formatting_binary(self, *args):
        print('Icon formatting (STEP1) ......')
        col_letter_dict = self._search_col_letter(*args)
        for column in args:            
            for sheet_num, sheet in enumerate(self.sheets):
                try:
                    col_letter = col_letter_dict[self.sheet_names[sheet_num]][column]
                except KeyError:
                    continue
                first = FormatObject(type='num', val=2, gte=True)
                mid = FormatObject(type='num', val=3, gte=True)
                last =FormatObject(type='num', val=4, gte=True)
                iconset = IconSet(iconSet='3Symbols2', cfvo=[first, mid, last], showValue=False, reverse=True)
                rule = Rule(type='iconSet', iconSet=iconset)
                formatting_region = f'{col_letter}1:{col_letter}{sheet.max_row}'
                sheet.conditional_formatting.add(formatting_region, rule)

    def formatting_tertiary(self, *args):
        print('Icon formatting (STEP2) ......')
        col_letter_dict = self._search_col_letter(*args)
        for column in args:            
            for sheet_num, sheet in enumerate(self.sheets):
                try:
                    col_letter = col_letter_dict[self.sheet_names[sheet_num]][column]
                except KeyError:
                    continue

                first = FormatObject(type='num', val=2, gte=True)
                mid = FormatObject(type='num', val=3, gte=True)
                last =FormatObject(type='num', val=4, gte=True)
                iconset = IconSet(iconSet='3Symbols2', cfvo=[first, mid, last], showValue=False, reverse=True)
                rule = Rule(type='iconSet', iconSet=iconset)
                
                formatting_region = f'{col_letter}1:{col_letter}{sheet.max_row}'
                sheet.conditional_formatting.add(formatting_region, rule)

    def overall_font_setting(self, font, size):
        overall_font = Font(name=font, size=size)
        for sheet_num, sheet in enumerate(self.sheets):
            for row in sheet:
                for cell in row:
                    sheet[cell.coordinate].font = overall_font

    def _save(self):
        input_path = pathlib.Path(self.input_file)
        input_file_name = input_path.stem
        input_file_dir = input_path.parent
        output = f'{input_file_dir}/{input_file_name}.formatted.xlsx'
        print(f'Creating {output} ......')
        self.workbook.save(output)

    # def formatting_quaternary(worksheet, column):
    #     first = FormatObject(type='num', val=1, gte=True)
    #     second = FormatObject(type='num', val=2, gte=True)
    #     third = FormatObject(type='num', val=3, gte=True)
    #     last =FormatObject(type='num', val=4, gte=True)
    #     iconset = IconSet(iconSet='4Arrows', cfvo=[first, second, third, last], showValue=True, reverse=True)
    #     rule = Rule(type='iconSet', iconSet=iconset)
    #     worksheet.conditional_formatting.add(column_regions_dic[column], rule)